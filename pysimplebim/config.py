from openpyxl.styles import PatternFill

class Config():
    validation_flag = True
    row_span = 100
    # Choose between these colors, and add more if you like
    colors = {

        "light_green": PatternFill("solid",start_color="AAFF80"),
        "dark_green": PatternFill("solid",start_color="009933"),
        "green": PatternFill("solid",start_color="009933"),
        "yellow": PatternFill("solid",start_color="FFFF00"),
        "orange": PatternFill("solid",start_color="FF9900"),
        "light_blue": PatternFill("solid",start_color="80DFFF"),
        "red": PatternFill("solid",start_color="CD5C5C"),
        "blue": PatternFill("solid",start_color="00CCFF"),
        "dark_blue": PatternFill("solid",start_color="0000E6"),
        "light_grey": PatternFill("solid",start_color="E6E6E6"),
        "grey": PatternFill("solid",start_color="999999"),
        "dark_grey": PatternFill("solid",start_color="595959"),
        "pink": PatternFill("solid",start_color="FF80FF"),
        "purple": PatternFill("solid",start_color="CC00FF"),
        "black": PatternFill("solid",start_color="1A1A1A"),
        "white": PatternFill("solid",start_color="FFFFFF"),
        "mmi_100": PatternFill("solid",start_color="EE2D28"),
        "mmi_200": PatternFill("solid",start_color="F3B72D"),
        "mmi_300": PatternFill("solid",start_color="E6E64B"),
        "mmi_350": PatternFill("solid",start_color="81C563"),
        "mmi_400": PatternFill("solid",start_color="2BB05D"),
        "mmi_500": PatternFill("solid",start_color="0F75BC")
    }
    ifc_import_settings = {
        "import_2d_annotations": None,
        "import_3d_faces": None
    }
    ifc_export_settings = {
        "ifc_file_mark": None,
        "export_space_boundaries": None,
        "export_wall_to_wall_connections": None,
        "export_grids": None,
        "export_color_of_3d_objects": "Yes",
        "export_legacy_colors_of_3d_objects": "Yes",
        "export_3d_objects_with_no_or_invalid_geometry": None,
        "exclude_openings_of_excluded_3d_objects": None,
        "exclude_all_openings": None,
        "optimize_shared_geometries": None,
        "export_model_information": None
    }
