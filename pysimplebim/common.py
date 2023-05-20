from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment
import inspect

from .config import Config
from .constraints import Dropdowns as d 


class Common():
    """Class containing all common functions"""
    def __init__(self, workbook): 
        self.wb = load_workbook(workbook)
        self.checking_span = 100
        self.first_column = 3
        self.no_errors = True

        # Variables for storing in Settings-sheet
        # TODO (Maybe)


    def find_functions(self, worksheet):
        """Updates dict containing all function names and first cell in function.
            Then Updates list with all function names"""
        self.functions = []
        self.function_coordinates = {}
        for y in range(1,1000):
            if (worksheet.cell(column=2, row=y).value == "*"):
                self.function_coordinates[worksheet[f"B{y-1}"].value] = f"C{y+1}"
                self.functions.append(worksheet[f"B{y-1}"].value)
        return None


    def find_function_start(self, function):
        """Returns first empty cell in a SB-function in excel-template"""
        first_cell = self.ws[self.function_coordinates[function]]
        i = first_cell.row

        current_cell = first_cell
        for i in range(i, i+self.checking_span):
            if self.ws.cell(row=i, column =self.first_column).value == None:
                current_cell = self.ws.cell(row=i, column=self.first_column)
                break
        return current_cell


    def increment_cell(self, current_cell):
        """Returns the right-neighbour-cell of the given cell"""
        current_row = current_cell.row
        current_column = current_cell.column
        current_cell = self.ws.cell(row = current_row, column=current_column+1)
        return current_cell
    

    def insert_value_in_cell(self, current_cell, value):
        """Insert the given value in the given cell"""
        current_cell.value = value
        return current_cell

    
    def insert_row_after(self, current_cell, num_funcion_parameters):
        """Insert an empty row below current cell. Also emulate border settings"""
        
        self.ws.insert_rows(current_cell.row + 1)
        new_row = current_cell.row+1
        bd_thin = Side(border_style="thin", color="000000")
        bd_thick= Side(border_style="thick", color="000000")
        start_column = self.first_column
        i=0 
        while i<num_funcion_parameters:
            current_cell = self.ws.cell(row=new_row, column=start_column+i)
            if i == num_funcion_parameters-1:
                current_cell.border = Border(left=bd_thin, right=bd_thick, bottom=bd_thin)
            else:
                current_cell.border = Border(left=bd_thin,bottom=bd_thin)
            i +=1
        return None
        

    def get_function_parameters(self, function):
        """Return a formatted list of all parameters in the given function"""
        
        signatures = str(inspect.signature(function))
        excess_characters = ["(", ")", "str", "=", "None", ":", " "] 
        for i in excess_characters:
            signatures= signatures.replace(i, "")
        signatures= signatures.split(",")
        return signatures
    

    def handle_list(self, argument):
        """If the argument is a list: Return the list as a newline rawstring. Each list element is seperated by a newline."""
        if type(argument) == list:
            return "\n".join(argument)
            

    def set_textWrap(self, cell):  
        """Set text_wrap=True for current cell. This is handy when working with lists"""
        cell.alignment = Alignment(wrap_text=True)  
        return None

    def set_cell_color(self, cell, color):
        """Set fill-color in given cell"""
        try:
            cell.fill = Config.colors[color]
        except:
            print("Please choose one of the following valid colors:")
            for key in Config.colors.keys():
                print(key)
        return None


    def write_row(self, params, locals, cell, dictionary):
        """Writes the given arguments to cells in a row"""
        i = 0
        while i<len(params) and self.no_errors:
            # Route to own function if input is a list
            if type(locals[params[i]]) == list:
                cell.value = self.handle_list(locals[params[i]])
                self.set_textWrap(cell)
            elif params[i] == "color":
                self.set_cell_color(cell, locals[params[i]])
            else:
                value = (locals[params[i]])
                parameter = params[i]
                
                # Validate that values that are predefined. Validation can be turned off in the Config-class, by setting the value to 
                if Config.validation_flag:
                    self.validate_cell(value, parameter, dictionary, i)
                # If validation is OK, assign value to cell
                cell.value = value

            cell = self.increment_cell(cell)
            i+=1
        return None
    

    def validate_cell(self, value, parameter, dictionary, i):
        """ Give the user a warning if the value for a given cell is not valid.
        """
        # Check that required values are filled out
        parameter_name = dictionary[i]["name"]
        function_name = dictionary["function_name"]
        if (value == None) and (dictionary[i]["required"] == True):          
            print(f"'{parameter_name}' is required in function {function_name} but no value vas given.\nPlease give an entry")
            self.no_errors = False
        # Check that predefined values are correct
        elif dictionary[i]["predefined"] == True and value != None:
            if value not in dictionary[i]["source"]:
                print (f"Please give a valied entry for parameter '{parameter_name}' in function '{function_name}'. \nAllowed values are:")
                for element in dictionary[i]["source"]:
                    print(f"\t{element}")
                self.no_errors = False      
        return None
          

    def set_active_worksheet(self, worksheet="Settings"):
        """ What are we doing here?"""
        if worksheet not in d.worksheets:
            print(f"{worksheet} is not a valid entry")
            print("Please choose from one of the following:")
            for i in d.worksheets:
                print(i)
        else:
            self.wb.active = self.wb[f"{worksheet}"]
            self.ws = self.wb.active


    def save_workbook(self,path=str):
        """Save the .xlsx-file to disk"""
        self.wb.save(path)
        return None
    
    def list_functions(self):
        """Prints all available funtions in the worksheet"""
        function_list = []
        for i in self.functions:
            print(i)
            function_list.append(i)
        return function_list


    def list_function_parameters(function_name):
        pass 

